import openpyxl
import os
import textfsm
from tkinter import Tk
from tkinter.filedialog import askdirectory

def main():
    Tk().withdraw()
    input_folder = askdirectory()

    output_files=os.scandir(input_folder)

    excel=openpyxl.Workbook()
    sheet=excel.create_sheet("Active Port check")

    sheet.cell(1,1).value="Hostname"
    sheet.cell(1,2).value="Interface"
    sheet.cell(1,3).value="Input packets"
    sheet.cell(1,4).value="Output packets"
    sheet.cell(1,5).value="Access"
    sheet.cell(1,6).value="Status"
    sheet.cell(1,7).value="Interface counter >0"
    sheet.cell(1,8).value="Cable connected"
    sheet.cell(1,9).value="10Mbps port"

    version_template= open(os.path.join("Templates","show_version.textfsm"))
    interface_template= open(os.path.join("Templates","show_interfaces.textfsm"))
    
    row=2
    for device in output_files:
        if "json" in device.name:
            continue

        print("Parsing: "+device.name)

        f = open(os.path.join(input_folder,device.name))
        cmd_output=f.read()
        f.close()

        template = textfsm.TextFSM(version_template)
        ver_fsm_results = template.ParseText(cmd_output)
        try:
            hostname=ver_fsm_results[0][2]
        except:
            print("Above device skipped due to no output.")
            hostname="NULL"
            #continue

        row,IP_list = scan_L2(sheet, cmd_output, interface_template, hostname, row)

        row+=1
    
    result_folder=str(input_folder.split("/")[-1])+"_Report"
    if not os.path.isdir(result_folder):
        os.mkdir(result_folder)

    excel.remove(excel["Sheet"])
    excel.save(os.path.join(result_folder,"L2_Scan.xlsx"))

def scan_L2(sheet, cmd_output, interface_template, hostname, row):
    template = textfsm.TextFSM(interface_template)
    fsm_results = template.ParseText(cmd_output)
    IP_list={}
    for item in fsm_results:
        sheet.cell(row,1).value=hostname
        sheet.cell(row,2).value=item[0]
        sheet.cell(row,3).value=item[21]
        sheet.cell(row,4).value=item[22]
        #sheet.cell(row,5)=
        if ("up" in item[1]) and ("up" in item[2]):
            sheet.cell(row,6).value="Yes"
        else:
            sheet.cell(row,6).value="No"
        try:
            if (int(item[21])>0 or int(item[22])>0):
                sheet.cell(row,7).value="Yes"
            else:
                sheet.cell(row,7).value="No"
        except:
            sheet.cell(row,7).value="No.?"
        if "10Mb/s" in item[10]:
            sheet.cell(row,9).value= "Yes"

        if "Vlan" in item[0]:
            IP_list[item[0]]=item[7]
        row+=1
    
    return row, IP_list


if __name__ == "__main__":
    main()