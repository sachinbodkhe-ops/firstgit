import openpyxl
from operator import itemgetter
import os
import textfsm
from tkinter import Tk
from tkinter.filedialog import askdirectory

def main():
    Tk().withdraw()
    input_folder = askdirectory()

    excel=openpyxl.Workbook()
    summary=excel.create_sheet("Show_CDP")

    output_files=os.scandir(input_folder)
    cdp_template= open(os.path.join("Templates","show_cdp_neighbors.textfsm"))
    version_template= open(os.path.join("Templates","show_version.textfsm"))
    inventory_template= open(os.path.join("Templates","show_inventory.textfsm"))

    row=2
    for device in output_files:
        if "json" in device.name:
            continue

        print("Parsing: "+device.name)

        f = open(os.path.join(input_folder,device.name))
        cmd_output=f.read()
        f.close()

        template = textfsm.TextFSM(version_template)
        ver_fsm_results = template.ParseText(cmd_output)

        try:
            hostname=ver_fsm_results[0][2]
        except:
            print("Above device skipped due to no output.")
            continue
        
        IP=str(device.name).split(".", 4)
        IP=".".join(IP[:4])

        row = scan_cdp(summary, cmd_output, cdp_template, inventory_template, hostname, IP, row)

        row+=1

    result_folder=str(input_folder.split("/")[-1])+"_Report"
    if not os.path.isdir(result_folder):
        os.mkdir(result_folder)

    excel.remove(excel["Sheet"])
    excel.save(os.path.join(result_folder,"CDP_Scan.xlsx"))
        
def scan_cdp(summary, cmd_output, cdp_template, inventory_template, hostname, IP, row):
    template = textfsm.TextFSM(cdp_template)
    cdp_fsm_results = template.ParseText(cmd_output)
    
    neighbors=[]    
    for item in cdp_fsm_results:
        capability=item[2]
        if ("R" in capability) or ("T" in capability) or ("B" in capability) or ("S" in capability) or ("I" in capability) or ("r" in capability):
            neighbors.append(item)
        #neighbors.append(item)

    neighbors=sorted(neighbors, key=itemgetter(1))

    summary.cell(row,2).value=hostname
    summary.cell(row,3).value=IP
    
    template = textfsm.TextFSM(inventory_template)
    inv_fsm_results = template.ParseText(cmd_output)
    
    SFP_info={}
    for item in inv_fsm_results:
        if "Ethernet" in item[0]:
            if "FastEthernet" in item[0]:
                item[0]=item[0].replace("FastEthernet", "Fas ")
            elif "TenGigabitEthernet" in item[0]:
                item[0]=item[0].replace("TenGigabitEthernet", "Ten ")
            elif "TwoGigabitEthernet" in item[0]:
                item[0]=item[0].replace("TwoGigabitEthernet", "Two ")
            elif "TwentyFiveGigE" in item[0]:
                item[0]=item[0].replace("TwentyFiveGigE", "Twe ")
            elif "FortyGigabitEthernet" in item[0]:
                item[0]=item[0].replace("FortyGigabitEthernet", "For ")
            elif "GigabitEthernet" in item[0]:
                item[0]=item[0].replace("GigabitEthernet", "Gig ")
            SFP_info[item[0]]=item[1]
        if ("Te" == item[0][:2]) and (len(item[0]) <= 18):
            item[0]=item[0].replace("Te","Ten ")
            SFP_info[item[0]]=item[1]
        elif ("Fo" == item[0][:2]) and (len(item[0]) <= 18):
            item[0]=item[0].replace("Fo","For ")
            SFP_info[item[0]]=item[1]

    for item in neighbors:
        summary.cell(row,4).value=item[0]
        summary.cell(row,5).value=item[1]
        summary.cell(row,6).value=item[2]
        summary.cell(row,7).value=item[3]
        summary.cell(row,8).value=item[4]
        if item[1] in SFP_info.keys():
            summary.cell(row,9).value=SFP_info[item[1]]
        row+=1
    
    if len(neighbors) < 1:
        return row+1
    return row

if __name__ == "__main__":
    main()