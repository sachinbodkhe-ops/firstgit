import json
import openpyxl
from operator import itemgetter
import os
import textfsm
from tkinter import Tk
from tkinter.filedialog import askdirectory

def main():
    Tk().withdraw()
    input_folder = askdirectory()

    excel=openpyxl.Workbook()
    sheet=excel.create_sheet("Endpoint_Discovery")

    sheet.cell(1,1).value="Switch"
    sheet.cell(1,2).value="MAC_Address"
    sheet.cell(1,3).value="Interface"
    sheet.cell(1,4).value="Interface Speed"
    sheet.cell(1,5).value="Interface Type"
    sheet.cell(1,6).value="VLAN"
    sheet.cell(1,7).value="IP"
    sheet.cell(1,8).value="Vendor"
    sheet.cell(1,9).value="CDP_Platform"
    sheet.cell(1,10).value="CDP_Hostname"
    sheet.cell(1,11).value="STATIC IP or DHCP"
    sheet.cell(1,11).value="Cisco Comments"
    sheet.cell(1,11).value="Customer Comments"
    sheet.cell(1,11).value="Fabric IP (Old/New/Remove)"

    with open(os.path.join("Templates","Mac_vendor_list.json"),"r") as f:
        mac_DB=json.load(f)

    SVI_list,ARP_Table= scan_SVI_and_arp(input_folder)

    output_files=os.scandir(input_folder)
    version_template= open(os.path.join("Templates","show_version.textfsm"))
    mac_tempalte= open(os.path.join("Templates","show_mac_address-table.textfsm"))
    cdp_template= open(os.path.join("Templates","show_cdp_neighbors.textfsm"))
    intf_template= open(os.path.join("Templates","show_interfaces_status.textfsm"))

    row=2
    for device in output_files:
        if "json" in device.name:
            continue

        print("Parsing: "+device.name)

        f = open(os.path.join(input_folder,device.name))
        cmd_output=f.read()
        f.close()

        template = textfsm.TextFSM(version_template)
        fsm_results = template.ParseText(cmd_output)
        
        try:
            hostname=fsm_results[0][2]
        except:
            print("Above device skipped due to no output.")
            continue

        row= endpoint_discovery(sheet, cmd_output, mac_tempalte, cdp_template, hostname, ARP_Table, mac_DB, intf_template, row)

    result_folder=str(input_folder.split("/")[-1])+"_Report"
    if not os.path.isdir(result_folder):
        os.mkdir(result_folder)

    excel.remove(excel["Sheet"])
    excel.save(os.path.join(result_folder,"Discovery.xlsx"))


def scan_SVI_and_arp(input_folder):
    output_files=os.scandir(input_folder)
    intf_template= open(os.path.join("Templates","show_ip_interface_brief.textfsm"))
    arp_template= open(os.path.join("Templates","show_arp.textfsm"))
    
    SVI_list={}
    ARP_Table={}

    for device in output_files:
        f = open(os.path.join(input_folder,device.name))
        cmd_output=f.read()
        f.close()
        
        template = textfsm.TextFSM(intf_template)
        fsm_results = template.ParseText(cmd_output)

        for interface in fsm_results:
            if "Vlan" in interface[0]:
                if interface[0] in SVI_list.keys():
                    SVI_list[interface[0]].append(device.name)
                else:
                    SVI_list[interface[0]]=list()
                    SVI_list[interface[0]].append(device.name)

        template = textfsm.TextFSM(arp_template)
        table = template.ParseText(cmd_output)
        
        for item in table:
            if item[2] in ARP_Table.keys():
                if "Vlan" in item[4]:
                    ARP_Table[item[2]]=[item[0],item[4]]
            else:
                ARP_Table[item[2]]=[item[0],item[4]]

    intf_template.close()
    arp_template.close()
    return SVI_list,ARP_Table


def endpoint_discovery(sheet, cmd_output, mac_tempalte, cdp_template, hostname, ARP_Table, mac_DB, intf_template, row):

        template = textfsm.TextFSM(mac_tempalte)
        mac_table = template.ParseText(cmd_output)

        for item in mac_table:
            if "Fa" == item[3][:2]:
                item[3]=item[3].replace("Fa","FastEthernet")
            elif "Twe" == item[3][:3]:
                item[3]=item[3].replace("Twe","TwentyFiveGigE")
            elif "Tw" == item[3][:2]:
                item[3]=item[3].replace("Tw","TwoGigabitEthernet")
            elif "Fo" == item[3][:2]:
                item[3]=item[3].replace("Fo","FortyGigabitEthernet")
            elif ("Gi" == item[3][:2]) and (len(item[3]) <= 15):
                item[3]=item[3].replace("Gi","GigabitEthernet")
            elif ("Te" == item[3][:2]) and (len(item[3]) <= 18):
                item[3]=item[3].replace("Te","TenGigabitEthernet")
        
        mac_table=sorted(mac_table, key=itemgetter(3))

        template = textfsm.TextFSM(cdp_template)
        cdp_neighbors = template.ParseText(cmd_output)

        for item in cdp_neighbors:
            if "Fas " == item[1][:4]:
                item[1]=item[1].replace("Fas ","FastEthernet")
            elif "Ten " == item[1][:4]:
                item[1]=item[1].replace("Ten ","TenGigabitEthernet")
            elif "Two " == item[1][:4]:
                item[1]=item[1].replace("Two ","TwoGigabitEthernet")
            elif "Twe " == item[1][:4]:
                item[1]=item[1].replace("Twe ","TwentyFiveGigE")
            elif "For " == item[1][:4]:
                item[1]=item[1].replace("For ","FortyGigabitEthernet")
            elif "Gig " == item[1][:4]:
                item[1]=item[1].replace("Gig ","GigabitEthernet")
        
        cdp_result={}
        for item in cdp_neighbors:
            if item[1] in cdp_result.keys():
                cdp_result[item[1]].append([item[0],item[3],item[2]])
            else:
                cdp_result[item[1]]=[[item[0],item[3],item[2]]]
        
        template = textfsm.TextFSM(intf_template)
        fsm_results = template.ParseText(cmd_output)

        for item in fsm_results:
            if "Fa" == item[0][:2]:
                item[0]=item[0].replace("Fa","FastEthernet")
            elif "Twe" == item[0][:3]:
                item[0]=item[0].replace("Twe","TwentyFiveGigE")
            elif "Tw" == item[0][:2]:
                item[0]=item[0].replace("Tw","TwoGigabitEthernet")
            elif "Fo" == item[0][:2]:
                item[0]=item[0].replace("Fo","FortyGigabitEthernet")
            elif "Gi" == item[0][:2]:
                item[0]=item[0].replace("Gi","GigabitEthernet")
            elif ("Te" == item[0][:2]) and (len(item[3]) <= 18):
                item[0]=item[0].replace("Te","TenGigabitEthernet")

        intf_type_speed={}
        for item in fsm_results:
            ver=list()
            try:
                int_type=int(item[3])
                ver.append("access")
            except:
                ver.append(item[3])
            ver.append(item[5])
            intf_type_speed[item[0]]=ver

        for mac in mac_table:
            if ("CPU" in mac[3]) or ("Drop" in mac[3]) or ("Vl" in mac[3][:2]) or ("Po" in mac[3][:2]) or ("Switch" in mac[3]):
                continue
            
            try:
                cdp_entry=cdp_result[mac[3]]
            except:
                cdp_entry=[]
            cdp_hostname=""
            cdp_platform=""
            #flag=False
            flag_2=False
            delimiter=""
            for item in cdp_entry:
                #if ("R" in item[2]) or ("S" in item[2]):
                #    flag=True
                #    break
                if flag_2:
                    delimiter=", "
                cdp_hostname=cdp_hostname+delimiter+item[0]
                cdp_platform=cdp_platform+delimiter+item[1]
                flag_2=True
            #if flag:
            #    continue

            sheet.cell(row,1).value=hostname
            sheet.cell(row,2).value=mac[0]
            sheet.cell(row,3).value=mac[3]
            try:
                sheet.cell(row,4).value=intf_type_speed[mac[3]][1]
                sheet.cell(row,5).value=intf_type_speed[mac[3]][0]
            except:
                pass
            sheet.cell(row,6).value=mac[2]

            svi="Vlan"+mac[2]
            try:
                if (mac[0] in ARP_Table.keys()) and (ARP_Table[mac[0]][1] == svi):
                    sheet.cell(row,7).value=ARP_Table[mac[0]][0]
            except:
                pass
            
            oui=mac[0].replace(".","")[:6]
            try:
                vendor=mac_DB[oui]
            except:
                vendor="Unknown"
            sheet.cell(row,8).value=vendor

            sheet.cell(row,9).value=cdp_platform
            sheet.cell(row,10).value=cdp_hostname
            
            row+=1
        return row

if __name__ == "__main__":
    main()