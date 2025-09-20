import json
import openpyxl
from operator import itemgetter
import os
import textfsm
from tkinter import Tk
from tkinter.filedialog import askdirectory

import scan_cdp
import scan_interfaces_L2
import scan_VLANs
import scan_show_run_for_vlan
import endpoint_discovery

def main():
    Tk().withdraw()
    input_folder = askdirectory()
    
    excel=openpyxl.Workbook()
    cdp_sheet=excel.create_sheet("CDP Scan")

    l2_sheet=excel.create_sheet("Active Port check")

    l2_sheet.cell(1,1).value="Hostname"
    l2_sheet.cell(1,2).value="Interface"
    l2_sheet.cell(1,3).value="Input packets"
    l2_sheet.cell(1,4).value="Output packets"
    l2_sheet.cell(1,5).value="Access"
    l2_sheet.cell(1,6).value="Status"
    l2_sheet.cell(1,7).value="Interface counter >0"
    l2_sheet.cell(1,8).value="Cable connected"
    l2_sheet.cell(1,9).value="10Mbps port"

    vlan_sheet=excel.create_sheet("VLAN Scan")
    data_voice_vlan_sheet=excel.create_sheet("Data and Voice VLANs")
    endp_sheet=excel.create_sheet("Endpoint Discovery")

    endp_sheet.cell(1,1).value="Switch"
    endp_sheet.cell(1,2).value="MAC_Address"
    endp_sheet.cell(1,3).value="Interface"
    endp_sheet.cell(1,4).value="Interface Speed"
    endp_sheet.cell(1,5).value="Interface Type"
    endp_sheet.cell(1,6).value="VLAN"
    endp_sheet.cell(1,7).value="IP"
    endp_sheet.cell(1,8).value="Vendor"
    endp_sheet.cell(1,9).value="CDP_Platform"
    endp_sheet.cell(1,10).value="CDP_Hostname"

    cdp_template= open(os.path.join("Templates","show_cdp_neighbors.textfsm"))
    version_template= open(os.path.join("Templates","show_version.textfsm"))
    inventory_template= open(os.path.join("Templates","show_inventory.textfsm"))
    interface_template= open(os.path.join("Templates","show_interfaces.textfsm"))
    vlan_template= open(os.path.join("Templates","show_vlan.textfsm"))
    ip_intf_template= open(os.path.join("Templates","show_ip_interface_brief.textfsm"))
    intf_desc_tempalte= open(os.path.join("Templates","show_interfaces_description.textfsm"))
    mac_tempalte= open(os.path.join("Templates","show_mac_address-table.textfsm"))
    intf_template= open(os.path.join("Templates","show_interfaces_status.textfsm"))

    with open(os.path.join("Templates","Mac_vendor_list.json"),"r") as f:
        mac_DB=json.load(f)

    output_files=os.scandir(input_folder)
    SVI_list,ARP_Table= endpoint_discovery.scan_SVI_and_arp(input_folder)

    output_files=os.scandir(input_folder)

    cdp_row=2
    l2_row=2
    vlan_row=2
    d_v_vlan_row=3
    endpoint_discovery_row=2
    for device in output_files:
        if "json" in device.name:
            continue

        print("Parsing: "+device.name)
        
        f = open(os.path.join(input_folder,device.name))
        cmd_output=f.read()
        f.close()

        template = textfsm.TextFSM(version_template)
        ver_fsm_results = template.ParseText(cmd_output)

        try:
            hostname=ver_fsm_results[0][2]
        except:
            print("Above device skipped due to no output.")
            continue

        IP=str(device.name).split(".", 4)
        IP=".".join(IP[:4])

        cdp_row = scan_cdp.scan_cdp(cdp_sheet, cmd_output, cdp_template, inventory_template, hostname, IP, cdp_row)
        cdp_row+=1

        l2_row,IP_list = scan_interfaces_L2.scan_L2(l2_sheet, cmd_output, interface_template, hostname, l2_row)

        vlan_row= scan_VLANs.scan_VLAN(vlan_sheet, cmd_output, vlan_template, ip_intf_template, intf_desc_tempalte, hostname, IP, vlan_row, IP_list)
        
        d_v_vlan_row= scan_show_run_for_vlan.scan_VLAN_2(data_voice_vlan_sheet, cmd_output, hostname, IP, d_v_vlan_row)

        endpoint_discovery_row= endpoint_discovery.endpoint_discovery(endp_sheet, cmd_output, mac_tempalte, cdp_template, hostname, ARP_Table, mac_DB, intf_template, endpoint_discovery_row)

    file_name=str(input_folder.split("/")[-1])+"_Report"
    if not os.path.isdir("Reports"):
        os.mkdir("Reports")


    excel.remove(excel["Sheet"])
    excel.save(os.path.join("Reports",file_name+".xlsx"))

if __name__ == "__main__":
    main()