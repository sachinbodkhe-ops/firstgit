import openpyxl
import os
from tkinter import Tk
from tkinter.filedialog import askdirectory
import re

def main():
    Tk().withdraw()
    input_folder = askdirectory()

    output_files=os.scandir(input_folder)

    excel=openpyxl.Workbook()
    sheet=excel.create_sheet("Scan_of_show_run")

    row=3
    for device in output_files:
        if "json" in device.name:
            continue

        print("Parsing: "+device.name)
        
        f = open(os.path.join(input_folder,device.name))
        cmd_output=f.read()
        f.close()

        try:
            hostname= re.search("hostname\s+(.+)",cmd_output).group(1)
        except:
            print("Above device skipped due to no output.")
            continue

        IP=str(device.name).split(".", 4)
        IP=".".join(IP[:4])

        row= scan_VLAN_2(sheet, cmd_output, hostname, IP, row)
    
    result_folder=str(input_folder.split("/")[-1])+"_Report"
    if not os.path.isdir(result_folder):
        os.mkdir(result_folder)

    excel.remove(excel["Sheet"])
    excel.save(os.path.join(result_folder,"Show_Run_VLAN_Scan.xlsx"))

def scan_VLAN_2(sheet, cmd_output, hostname, IP, row):
    sheet.cell(row,2).value=hostname
    sheet.cell(row,3).value=IP
    
    vlans=re.findall("switchport access vlan (\d+)",cmd_output)
    access_vlans={}
    for vlan in vlans:
        if vlan in access_vlans.keys():
            access_vlans[vlan]+=1
        else:
            access_vlans[vlan]=1

    vlans=re.findall("switchport voice vlan (\d+)",cmd_output)
    voice_vlans={}
    for vlan in vlans:
        if vlan in voice_vlans.keys():
            voice_vlans[vlan]+=1
        else:
            voice_vlans[vlan]=1
    
    sheet.cell(row,4).value="Access VLANs"
    sheet.cell(row,5).value="No of Ports"
    row_1=row+1
    for key,val in access_vlans.items():
        sheet.cell(row_1,4).value=key
        sheet.cell(row_1,5).value=val
        row_1+=1

    sheet.cell(row,7).value="Voice VLANs"
    sheet.cell(row,8).value="No of Ports"
    row_2=row+1
    for key,val in voice_vlans.items():
        sheet.cell(row_2,7).value=key
        sheet.cell(row_2,8).value=val
        row_2+=1

    if row_1 >= row_2:
        return row_1+2
    else:
        return row_2+2

if __name__ == "__main__":
    main()