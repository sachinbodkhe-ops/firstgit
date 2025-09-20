import openpyxl
import os
import textfsm
from tkinter import Tk
from tkinter.filedialog import askdirectory

def main():
    Tk().withdraw()
    input_folder = askdirectory()

    output_files=os.scandir(input_folder)

    excel=openpyxl.Workbook()
    sheet=excel.create_sheet("Scan_of_VLANs")

    sheet.cell(1,2).value="Hostname"
    sheet.cell(1,3).value="IP"

    version_template= open(os.path.join("Templates","show_version.textfsm"))
    vlan_template= open(os.path.join("Templates","show_vlan.textfsm"))
    ip_intf_template= open(os.path.join("Templates","show_ip_interface_brief.textfsm"))
    intf_desc_tempalte= open(os.path.join("Templates","show_interfaces_description.textfsm"))

    row=3
    for device in output_files:
        if "json" in device.name:
            continue

        print("Parsing: "+device.name)
        
        f = open(os.path.join(input_folder,device.name))
        cmd_output=f.read()
        f.close()

        template = textfsm.TextFSM(version_template)
        ver_fsm_results = template.ParseText(cmd_output)

        try:
            hostname=ver_fsm_results[0][2]
        except:
            print("Above device skipped due to no output.")
            continue

        IP=str(device.name).split(".", 4)
        IP=".".join(IP[:4])

        row= scan_VLAN(sheet, cmd_output, vlan_template, ip_intf_template, intf_desc_tempalte, hostname, IP, row)
    
    result_folder=str(input_folder.split("/")[-1])+"_Report"
    if not os.path.isdir(result_folder):
        os.mkdir(result_folder)

    excel.remove(excel["Sheet"])
    excel.save(os.path.join(result_folder,"VLAN_Scan.xlsx"))

def scan_VLAN(sheet, cmd_output, vlan_template, ip_intf_template, intf_desc_tempalte, hostname, IP, row, IP_list):

    sheet.cell(row,2).value=hostname
    sheet.cell(row,3).value=IP
    row+=1

    template = textfsm.TextFSM(vlan_template)
    fsm_results = template.ParseText(cmd_output)

    sheet.cell(row,4).value="VLAN"
    sheet.cell(row,5).value="Name"
    sheet.cell(row,6).value="Status"
    sheet.cell(row,7).value="No of Ports"
    row_1=row+1
    for vlan in fsm_results:
        if ("1002" in vlan[0]) or ("1003" in vlan[0]) or ("1004" in vlan[0]) or ("1005" in vlan[0]):
            continue
        sheet.cell(row_1,4).value=vlan[0]
        sheet.cell(row_1,5).value=vlan[1]
        sheet.cell(row_1,6).value=vlan[2]
        sheet.cell(row_1,7).value=len(vlan[3])
        row_1+=1

    template = textfsm.TextFSM(ip_intf_template)
    fsm_results = template.ParseText(cmd_output)

    interfaces={}
    for interface in fsm_results:
        if "Vlan" in interface[0]:
            values=[]
            values.append(interface[1])
            values.append(interface[2])
            values.append(interface[3])
            interfaces[interface[0]]=values

    template = textfsm.TextFSM(intf_desc_tempalte)
    fsm_results = template.ParseText(cmd_output)

    sheet.cell(row,9).value="SVI"
    sheet.cell(row,10).value="Description"
    sheet.cell(row,11).value="IP Address/Subnet"
    sheet.cell(row,12).value="Status"
    sheet.cell(row,13).value="Protocol"
    
    row_2=row+1
    for interface in fsm_results:
        if "Vl" in interface[0]:
            intf=interface[0]
            intf=intf.replace("Vl","Vlan")
            
            sheet.cell(row_2,9).value=intf
            sheet.cell(row_2,10).value=interface[3]
            try:
                if IP_list[intf]:
                    sheet.cell(row_2,11).value=IP_list[intf]
                else:
                    sheet.cell(row_2,11).value=interfaces[intf][0]
            except:
                sheet.cell(row_2,11).value=interfaces[intf][0]
            sheet.cell(row_2,12).value=interfaces[intf][1]
            sheet.cell(row_2,13).value=interfaces[intf][2]
            row_2+=1
    
    if row_1 >= row_2:
        return row_1+2
    else:
        return row_2+2

if __name__ == "__main__":
    main()